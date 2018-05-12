import requests
import argparse
from random import choices
from collections import defaultdict
from contextlib import contextmanager

from lxml import etree
from bs4 import BeautifulSoup
from openpyxl import Workbook


def fetch_url_of_each_course():
    site_url = "https://www.coursera.org/sitemap~www~courses.xml"
    response = requests.get(site_url)
    url_list = []
    first_children = 0
    tree = etree.fromstring(response.content)
    for sitemap in tree:
        children = sitemap.getchildren()
        url_list.append(children[first_children].text)
    return url_list


def fetch_html(url):
    response = requests.get(url)
    return response.text


def weeks_format(weeks):
    if weeks == 1:
        weeks = '1 week'
    else:
        weeks = ' '.join([str(weeks), 'weeks'])
    return weeks


def avarage_rating_format(rating):
    first_occurance = 0
    if rating:
        rating = rating[first_occurance].text
        trash_start = rating.find('See what')
        rating = rating[:trash_start]
    else:
        rating = 'There isn`t any user ratings'
    return rating


def parse_course_page(html_page):
    soup = BeautifulSoup(html_page, 'html.parser')
    first_occurance = 0
    course_name = soup.find_all('h1', class_='title display-3-text') \
	[first_occurance].text
    language = soup.select('div.language-info')[first_occurance].text
    weeks_amount = len(soup.find_all('div', class_='week'))
    start_date = list(soup.find_all('div', class_='startdate '
        'rc-StartDateString caption-text')[first_occurance].children) \
	[first_occurance].text
    avarage_rating = soup.find_all('div', class_='ratings-text bt3-hidden-xs')
    course_info = {
            'course_name': course_name,
            'language': language,
            'weeks_amount': weeks_format(weeks_amount),
            'start_date': start_date,
            'avg_rating': avarage_rating_format(avarage_rating)
    }
    return course_info

def parse_course_list(urls):
    courses_information = []
    for url in urls:
        html_page = fetch_html(url)
        course_info = parse_course_page(html_page)
        courses_information.append(course_info)
    return courses_information


@contextmanager
def open_workbook(workbook_name):
    workbook = Workbook()
    yield workbook
    workbook.save(''.join([workbook_name, '.xlsx']))


def fill_worksheet(ws, courses_info_list, title):
    ws.title = title
    for row_number, course_info in enumerate(courses_info_list, start=1):
        ws.cell(row=row_number, column=1, value=course_info['course_name'])
        ws.cell(row=row_number, column=2, value=course_info['language'])
        ws.cell(row=row_number, column=3, value=course_info['weeks_amount'])
        ws.cell(row=row_number, column=4, value=course_info['start_date'])
        ws.cell(row=row_number, column=5, value=course_info['avg_rating'])


def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument('-a', '--amount', default=5, type=int)
    parser.add_argument('-n', '--name', default='Coursera_courses', type=str)
    return parser.parse_args()


if __name__ == '__main__':
    args = parse_args()
    all_courses = fetch_url_of_each_course()
    random_urls = choices(all_courses, k=args.amount)
    try:
        courses_data = parse_course_list(random_urls)
    except requests.exceptions.RequestException as error:
        print("Sorry, can't establish internet connection, here's error "
                "message: {}".format(error))
    with open_workbook(args.name) as wb:
        fill_worksheet(wb.active, courses_data, title='Courses_list')
